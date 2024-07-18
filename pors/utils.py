import codecs
import csv
import io
import json
import re
from hashlib import sha256
from threading import Thread
from typing import Optional
from urllib.parse import urlunparse

import jdatetime
import pytz
import xlsxwriter
from django.conf import settings
from django.core.mail import EmailMessage
from django.db import connection
from django.db.models import QuerySet
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from persiantools.jdatetime import JalaliDate
from rest_framework import status
from rest_framework.request import Request
from rest_framework.response import Response

from . import models as m
from .messages import Message

HR_SCHEME = "http"
HR_HOST = "192.168.20.81"
HR_PORT = "14000"
SERVER_PORT = "26000"
HR_PROFILE_PATH = "/media/HR/PersonalPhoto/"

REPORT_DATE_TIME = {
    "bold": True,
    "font_name": "Tahoma",
    "font_size": 10,
    "bg_color": "#cce6ff",
    "border": 1,
}


def localnow() -> jdatetime.datetime:
    utc_now = jdatetime.datetime.now(tz=pytz.utc)
    local_timezone = pytz.timezone("Asia/Tehran")
    return utc_now.astimezone(local_timezone)


def get_str(date: jdatetime.date) -> str:
    """Converting a Jalali date object to a string
    yyyy/mm/dd
    """
    return date.strftime("%Y/%m/%d")


def first_and_last_day_date(month: int, year: int) -> tuple[str, str]:
    """This function returns a Jalali date object from the start and end
    dates provided
    """

    # Todo convert year and month to gro
    last_day_of_month = JalaliDate.days_in_month(month, year)
    first_day_date = get_str(jdatetime.date(year, month, 1))
    last_day_date = get_str(jdatetime.date(year, month, last_day_of_month))

    return first_day_date, last_day_date


def get_weekend_holidays(year: int, month: int) -> list[jdatetime.date]:
    """This function is responsible for calculating weekend holidays"""

    # Todo get last day of month in jalali date
    last_day_of_month = JalaliDate.days_in_month(month, year)
    holidays = []
    for daynum in range(1, last_day_of_month + 1):
        daynum_date = jdatetime.date(year, month, daynum)
        if daynum_date.weekday() in (6, 5):
            holidays.append(daynum_date.strftime("%Y/%m/%d").replace("-", "/"))
    return holidays


def get_current_date() -> tuple[int, int, int]:
    """Returning current date"""
    now = localnow()
    return now.year, now.month, now.day


def split_dates(dates, mode: str):
    """
    Splitting date and returning requested section based on mode.

    Args:
        dates: List of dates, or a single date to split.
        mode: The section, choose between `year`, `month` and `day`, `all`.

    Returns:
        List or single integer.
        int | list[int]
    """
    new_dates = []
    mode = mode.lower()

    if mode == "day":
        if not isinstance(dates, list):
            return int(dates.split("/")[2])
        for date in dates:
            new_dates.append(int(date.split("/")[2]))
        return new_dates
    elif mode == "month":
        if not isinstance(dates, list):
            return int(dates.split("/")[1])
        for date in dates:
            new_dates.append(int(date.split("/")[1]))
        return new_dates
    elif mode == "year":
        if not isinstance(dates, list):
            return int(dates.split("/")[0])
        for date in dates:
            new_dates.append(int(date.split("/")[0]))
        return new_dates
    elif mode == "all":
        if not isinstance(dates, list):
            return (
                int(dates.split("/")[0]),
                int(dates.split("/")[1]),
                int(dates.split("/")[2]),
            )
        for date in dates:
            new_dates.append(
                (
                    int(dates.split("/")[0]),
                    int(dates.split("/")[1]),
                    int(dates.split("/")[2]),
                )
            )
        return new_dates


def split_json_dates(dates: str) -> dict[str, str]:
    """
    Splitting a json list of dates and generating a dict with day number.
    This function assumes that you have a `day` key in your json data that
    contains a VALID date

    Args:
        dates: serialized (json) data contains a `date` key.

    Returns:
        Deserialized data, date key will contain the number of day only.
    """

    dates = json.loads(dates)
    for obj in dates:
        obj["day"] = int(obj["day"].split("/")[2])
    return dates


def validate_date(date: str) -> Optional[str]:
    """
    Validating date value and format.
    Replacing "-" with "/" if the value is valid.

    Example:
        "1402/00/01" = valid
        "1402/0/1" = invalid, month and day section must have 2 integers.

    Args:
        date: the date for validation.

    Returns:
        date | None: date value or None if it was invalid.
    """

    pattern = r"^\d{4}\/\d{2}\/\d{2}$"
    if not isinstance(date, str):
        return None
    if "-" in date:
        date = date.replace("-", "/")
    if re.match(pattern, date):
        return date
    else:
        return None


def execute_raw_sql_with_params(query: str, params: tuple) -> list:
    """
    Executing raw queries via context manager

    Args:
        query: the raw query
        params: parameters used in query, avoiding sql injections

    Returns:
        result: the data retrieved by query
    """
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return result


def report_generated_date_and_time():
    # wrting date and time that report generated
    report_gen_at = localnow().strftime("%Y%m%d %H:%M:%S")
    return f"این گزارش  در {report_gen_at} از سیستم دریافت شده است"


def queryset_to_xlsx_response_food_provider_ordering(
    queryset, persian_headers, providers
):
    # get the max col letter
    query_col_number = len(queryset[0])
    end_col = get_column_letter(query_col_number)

    # Create an in-memory output file for the new workbook.
    output = io.BytesIO()

    # Create a new Excel workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet()

    # Define a format for the table that sets the font.
    table_font = workbook.add_format({"font_name": "Tahoma", "font_size": 10})

    # Add a bold format for headers.
    header_format = workbook.add_format(
        {
            "bold": True,
            "font_name": "Tahoma",
            "font_size": 10,
            "bg_color": "#D7E4BC",  # Light green background for headers
            "border": 1,
        }
    )

    date_time_format = workbook.add_format(REPORT_DATE_TIME)

    # spilting qs to providers
    providers_qs = {}
    for provider in providers:
        providers_qs[provider] = queryset.filter(FoodProviderPersian=provider)

    msg = report_generated_date_and_time()
    worksheet.merge_range(f"A1:{end_col}1", msg, date_time_format)

    main_row_number = 1
    for provider, qs in providers_qs.items():
        merge_range = f"A{main_row_number + 1}:{end_col}{main_row_number + 1}"
        worksheet.merge_range(merge_range, provider, header_format)
        main_row_number += 1

        # Write Persian headers to the first row.
        for col_num, header in enumerate(persian_headers):
            worksheet.write(
                0 + main_row_number, col_num, header, header_format
            )
        main_row_number += 1

        col_widths = [len(header) for header in persian_headers]

        for row_num, obj in enumerate(qs):
            for col_num, data in enumerate(obj.values()):
                value = str(data)  # Ensure value is a string
                worksheet.write_string(
                    row_num + main_row_number, col_num, value, table_font
                )  # Write as text

                # update col width
                col_widths[col_num] = max(col_widths[col_num], len(value))
        main_row_number += len(qs) + 2

    # Adjust col len
    for col_num, width in enumerate(col_widths):
        worksheet.set_column(col_num, col_num, width + 2)

    # Close the workbook.
    workbook.close()

    # Rewind the buffer.
    output.seek(0)

    # Create an HTTP response with the appropriate content type and headers.
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=report.xlsx"

    return response


def queryset_to_xlsx_response(queryset, persian_headers):
    # get the max col letter
    query_col_number = len(queryset[0])
    end_col = get_column_letter(query_col_number)

    # Create an in-memory output file for the new workbook.
    output = io.BytesIO()

    # Create a new Excel workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet_BRF = workbook.add_worksheet("صبحانه")
    worksheet_LNC = workbook.add_worksheet("ناهار")

    # Define a format for the table that sets the font.
    table_font = workbook.add_format({"font_name": "Tahoma", "font_size": 10})

    # Add a bold format for headers.
    header_format = workbook.add_format(
        {
            "bold": True,
            "font_name": "Tahoma",
            "font_size": 10,
            "bg_color": "#D7E4BC",  # Light green background for headers
            "border": 1,
        }
    )

    date_time_format = workbook.add_format(REPORT_DATE_TIME)

    # wrting date and time that report generated
    msg = report_generated_date_and_time()

    worksheet_BRF.merge_range(f"A1:{end_col}1", msg, date_time_format)
    worksheet_LNC.merge_range(f"A1:{end_col}1", msg, date_time_format)

    # Write Persian headers to the first row.
    for col_num, header in enumerate(persian_headers):
        worksheet_BRF.write(1, col_num, header, header_format)
        worksheet_LNC.write(1, col_num, header, header_format)

    col_widths = [len(header) for header in persian_headers]

    # Write data to the sheet Breakfast.
    for row_num, obj in enumerate(
        queryset.filter(MealType=m.MealTypeChoices.BREAKFAST), start=2
    ):
        for col_num, data in enumerate(obj.values()):
            value = str(data)  # Ensure value is a string
            worksheet_BRF.write_string(
                row_num, col_num, value, table_font
            )  # Write as text

            # update col width
            col_widths[col_num] = max(col_widths[col_num], len(value))

    # Write data to the sheet Lunch.
    for row_num, obj in enumerate(
        queryset.filter(MealType=m.MealTypeChoices.LAUNCH), start=2
    ):
        for col_num, data in enumerate(obj.values()):
            value = str(data)  # Ensure value is a string
            worksheet_LNC.write_string(
                row_num, col_num, value, table_font
            )  # Write as text

            # update col width
            col_widths[col_num] = max(col_widths[col_num], len(value))

    # Adjust col len
    for col_num, width in enumerate(col_widths):
        worksheet_BRF.set_column(col_num, col_num, width + 2)
        worksheet_LNC.set_column(col_num, col_num, width + 2)

    # Close the workbook.
    workbook.close()

    # Rewind the buffer.
    output.seek(0)

    # Create an HTTP response with the appropriate content type and headers.
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=report.xlsx"

    return response


def generate_csv(queryset: QuerySet):
    """
    This function generates a dynamic csv content based on
        the queryset data argument.

    Warnings:
        Please note that you have to customize your queryset via filter, values
            and other stuffs before using this function.
        All fields and values on received queryset will use in csv.

    Args:
        queryset: The queryset that you want to generate csv from it.

    Returns:
        str: csv content that generated from queryset.
    """
    response = HttpResponse(content_type="text/csv")
    response.write(codecs.BOM_UTF8)

    writer = csv.writer(response)

    headers_appended = False

    for obj in queryset:
        if isinstance(obj, dict):
            data = obj.values()
            if not headers_appended:
                writer.writerow(obj.keys())
                headers_appended = True
            writer.writerow(data)
        else:
            keys = []
            values = []
            for field in obj._meta.fields:
                keys.append(field.name)
                values.append(getattr(obj, field.name))

            if not headers_appended:
                writer.writerow(keys)
                headers_appended = True
            writer.writerow(values)

    return response


def validate_request_based_on_schema(schema: dict, data: dict):
    """
    This function is responsible for validating request data based on the
        provided schema.
    Validation is checked by both checking parameter names
        as well as their types.

    Args:
        schema (dict): Your prefered schema which you want
            to receive from requets
        data (dict): The request data.

    """

    schema_params = set(schema.keys())
    data_params = set(data.keys())
    diffs = schema_params.difference(data_params)
    if diffs:
        raise ValueError(f"{diffs} parameter(s) must specified.")

    for param in schema_params:
        if not isinstance(data.get(param), type(schema.get(param))):
            raise ValueError(f"Invalid {param} value.")


def get_specific_deadline(
    weekday: int,
    meal_type: m.Item.MealTypeChoices = None,
    deadline: tuple = None,
):
    """
    Returning the submission's deadline based on the mealtype it has.
    Deadline is fetched from db based on the weekday.

    If meal_type parameter is not specified, all deadlines related to that
        weekday will get returned instead.
    Args:
        meal_type: The submission's deadline.
        weekday: Number of weekday (due to dynamic deadline logic).
        deadline: named tuple for deadline that has Days and Hour args.

    Returns:
        Dict of specific weekday deadlines | Days and hour value.
        Dict[str, namedtuple[Days, Hour]] | tuple[int, int]
    """

    if not meal_type and deadline:
        qs = m.Deadlines.objects.filter(WeekDay=weekday)
        deadlines = {}
        for row in qs:
            if row.MealType == m.MealTypeChoices.BREAKFAST:
                deadlines["breakfast"] = deadline(row.Days, row.Hour)
            else:
                deadlines["launch"] = deadline(row.Days, row.Hour)
        return deadlines

    qs = m.Deadlines.objects.get(MealType=meal_type, WeekDay=weekday)

    if deadline:
        return deadline(qs.Days, qs.Hour)

    return qs.Days, qs.Hour


def generate_token_hash(
    personnel: str, full_name: str, random_bit: int
) -> str:
    packed_args = (
        personnel.encode()
        + full_name.encode()
        + bytes(str(random_bit), "utf-8")
    )
    return sha256(packed_args).hexdigest()


def get_personnel_from_token(token: str):
    return m.User.objects.filter(Token=token, IsActive=True).first()


def create_jdate_object(date: str) -> jdatetime.date:
    """
    Creating Jalali Date object from provided date.

    Args:
        date (str): The date you want to create object from.

    Returns:
        jdatetime.date
    """

    year, month, day = split_dates(date, mode="all")
    return jdatetime.date(year, month, day)


# For type annotation only
MealTypeDeadlines = dict[int, tuple[int, int], dict[int, tuple[int, int]]]


def get_deadlines(
    deadline: tuple,
) -> tuple[MealTypeDeadlines, MealTypeDeadlines]:
    """
    Fetching deadlines from database and forming 2 dicts from the data.
    Dicts are formed based on the meal type, one for each type.
    Dicts keys are the weekday nums, so we have 7 keys for each dict.

    Returns:
        tuple[MealTypeDeadlines, MealTypeDeadlines]:
        - MealTypeDeadlines: The dict that contains the day and hour deadline
            for each weekday number.

    Examples:
        breakfast_deadline[0] = (1, 12)
        Here the [0] means the first day of week (Shanbe / Saturday),
        and (1) is the Days deadline, (12) is Hour.

    """

    deadlines = m.Deadlines.objects.all()
    breakfast_deadlines = {}
    launch_deadlines = {}

    for row in deadlines:
        if row.MealType == m.MealTypeChoices.BREAKFAST:
            breakfast_deadlines[row.WeekDay] = deadline(row.Days, row.Hour)
        else:
            launch_deadlines[row.WeekDay] = deadline(row.Days, row.Hour)

    return breakfast_deadlines, launch_deadlines


def fetch_available_location():
    """fetch available location (building and floors from HR)"""
    from .serializers import BuildingSerializer

    building_code_prefix = "Building_"
    qs = m.HR_constvalue.objects.filter(Code__startswith=building_code_prefix)
    if not qs:
        return {}

    buildings = []
    for building in qs:
        floors_qs = m.HR_constvalue.objects.filter(Parent_id=building.id)
        buildings.append(
            dict(
                code=building.Code,
                title=building.Caption,
                floors=floors_qs,
            )
        )
    return BuildingSerializer(buildings, many=True).data


def raise_report_notfound(message_obj: Message, request: Request):
    msg = "هیچ رکوردی بین بازه ارائه داده شده موجود نیست."
    error = "Queryset is empty!"
    message_obj.add_message(request, msg, Message.ERROR)
    return Response(
        {"messages": message_obj.messages(request), "errors": error},
        status.HTTP_404_NOT_FOUND,
    )


def add_mealtype_building(order_row, schema: dict):
    """
    Adding buildings info for each meal type in each order record
    of personnel calendar.
    """

    if (
        order_row["MealType"] == m.MealTypeChoices.LAUNCH
        and schema.get("launchDeliveryBuilding") is None
    ):
        schema["launchDeliveryBuilding"] = order_row["DeliveryBuilding"]
        schema["launchDeliveryFloor"] = order_row["DeliveryFloor"]
    elif (
        order_row["MealType"] == m.MealTypeChoices.BREAKFAST
        and schema.get("breakfastDeliveryBuilding") is None
    ):
        schema["breakfastDeliveryBuilding"] = order_row["DeliveryBuilding"]
        schema["breakfastDeliveryFloor"] = order_row["DeliveryFloor"]


def profile_url(filename):
    path = HR_PROFILE_PATH + filename
    return urlunparse((HR_SCHEME, f"{HR_HOST}:{HR_PORT}", path, "", "", ""))


def send_email_notif(
    subject: str,
    message: str,
    emails: list[str],
    reason: m.EmailReason,
    max_tries: int = 2,
):
    email_thread = Thread(
        target=_send_mail,
        args=(subject, message, emails, max_tries, reason),
    )
    email_thread.start()


def _send_mail(
    subject: str,
    message: str,
    emails: list[str],
    max_tries: int,
    reason=m.EmailReason,
):
    email = EmailMessage(
        subject=subject, body=message, to=emails, bcc=[settings.EMAIL_HOST_USER], from_email=settings.EMAIL_HOST_USER
    )
    email.content_subtype = "html"

    total_tries = 0
    success = 0
    while success == 0:
        try:
            success = email.send()
            m.ActionLog.objects.log(
                action_type=m.ActionLog.ActionTypeChoices.CREATE,
                log_msg=f"[{reason.value}], email notif for sent successfully",
            )
        except Exception as e:
            total_tries += 1

            if total_tries < max_tries:
                continue
            else:
                m.ActionLog.objects.log(
                    action_type=m.ActionLog.ActionTypeChoices.CREATE,
                    log_msg=f"[{reason.value}], An error occured while tried to send email "
                    f"notif\nError:{str(e)}",
                )
                break


def str_date_to_jdate(date: str) -> jdatetime.date:
    year, month, day = split_dates(date, mode="all")

    return jdatetime.date(year, month, day)